import psycopg2
import pandas as pd
from psycopg2 import Error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import numpy as np
import math
import os
import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload

# --- Google Drive Configuration ---
# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/drive.file"]


def authenticate_google_drive():
    """Authenticates with the Google Drive API and returns the service object."""
    creds = None

    # Get the absolute path of the directory containing the running script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    token_path = os.path.join(script_dir, "token.json")
    credential_path = os.path.join(script_dir, "credentials.json")

    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first time.
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # --- PATH FIX: Use absolute path to credentials.json ---
            print(f"Searching for credentials file at: {credential_path}")
            if not os.path.exists(credential_path):
                print("\nERROR: Could not find credentials.json at the specified path.")
                print(">>> Please CHECK AGAIN and ensure that:")
                print(
                    "1. The file is named exactly 'credentials.json' (not 'credentials.json.txt')."
                )
                print(
                    "2. The 'credentials.json' file is placed in the SAME FOLDER as this Python script.\n"
                )
                return None

            flow = InstalledAppFlow.from_client_secrets_file(credential_path, SCOPES)
            creds = flow.run_local_server(port=0)

        # Save the credentials for the next run
        with open(token_path, "w") as token:
            token.write(creds.to_json())

    try:
        service = build("drive", "v3", credentials=creds)
        print("Google Drive authentication successful.")
        return service
    except HttpError as error:
        print(f"An error occurred while building the Drive service: {error}")
        return None


def upload_to_drive(service, file_path, folder_id=None):
    """Uploads a file to Google Drive, optionally to a specific folder."""
    if not service:
        print("Upload failed: Google Drive service is not authenticated.")
        return

    file_name = os.path.basename(file_path)
    # Define file metadata, including the parent folder if an ID is provided
    file_metadata = {"name": file_name}
    if folder_id:
        file_metadata["parents"] = [folder_id]

    media = MediaFileUpload(
        file_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    try:
        file = (
            service.files()
            .create(body=file_metadata, media_body=media, fields="id, webViewLink")
            .execute()
        )
        print(f"File upload successful! File ID: {file.get('id')}")
        print(f"You can view the file at: {file.get('webViewLink')}")
    except HttpError as error:
        print(f"An error occurred during file upload: {error}")


def export_postgres_to_excel(db_params, query, output_file):
    """
    Export data from PostgreSQL to an Excel file with custom formatting.
    """

    connection = None
    cursor = None
    writer = None

    try:
        # Step 1: Connect to PostgreSQL
        print("Connecting to PostgreSQL...")
        connection = psycopg2.connect(**db_params)
        cursor = connection.cursor()

        # Step 2: Execute the query
        print("Executing query...")
        cursor.execute(query)

        # Step 3: Fetch column names and data
        columns = [desc[0] for desc in cursor.description]
        data = cursor.fetchall()

        # Step 4: Create a pandas DataFrame
        print("Loading data into DataFrame...")
        print(f"Columns retrieved: {columns}")
        df = pd.DataFrame(data, columns=columns)
        if df.empty:
            print("Warning: Query returned no data. Creating empty Excel file.")
            df = pd.DataFrame(columns=columns)

        # --- CLEAR values at DataFrame index 27 for all columns except the first one ---
        if 27 in df.index:
            for col in df.columns[1:]:
                df[col] = df[col].astype(object)
                df.at[27, col] = ""

        # --- CLEAR values from DataFrame at indices 26–31 for columns "Head" through "Miền Trung" ---
        cols_to_clear = ["Head", "Miền Bắc", "Miền Nam", "Miền Trung"]
        for idx in [26, 27, 28, 29, 30, 31]:
            if idx in df.index:
                for col in cols_to_clear:
                    if col in df.columns:
                        df.at[idx, col] = ""

        # --- If indices 28–30 exist and "TOTAL" (uppercase), divide by 100 ---
        if "TOTAL" in df.columns:
            for idx in [28, 29, 30]:
                if idx in df.index:
                    raw = pd.to_numeric(df.loc[idx, "TOTAL"], errors="coerce")
                    df.loc[idx, "TOTAL"] = raw / 100

        # Lists of columns for dividing and formatting
        columns_to_divide_existing = ["Head", "Miền Bắc", "Miền Nam", "Miền Trung"]
        columns_to_divide_new = [
            "Đông Bắc Bộ",
            "Tây Bắc Bộ",
            "ĐB Sông Hồng",
            "Bắc Trung Bộ",
            "Nam Trung Bộ",
            "Tây Nam Bộ",
            "Đông Nam Bộ",
            "TOTAL",
        ]
        columns_to_divide_index_31 = (
            columns_to_divide_existing + ["Total"] + columns_to_divide_new
        )

        # === Step 4.1: Data transformations in DataFrame ===
        mask_existing = ~df.index.isin(range(28, 32))
        for col in columns_to_divide_existing:
            if col in df.columns:
                df.loc[mask_existing, col] = (
                    pd.to_numeric(df.loc[mask_existing, col], errors="coerce")
                    / 1_000_000
                ).round(2)
            else:
                print(f"Warning: '{col}' column not found.")

        if "Total" in df.columns:
            mask_total = ~df.index.isin([26] + list(range(28, 32)))
            df.loc[mask_total, "Total"] = (
                pd.to_numeric(df.loc[mask_total, "Total"], errors="coerce") / 1_000_000
            ).round(2)
        else:
            print("Warning: 'Total' column not found.")

        rows_to_divide_new_df_indices = [
            i for i in range(len(df)) if i not in [26, 28, 29, 30, 31]
        ]
        for col in columns_to_divide_new:
            if col in df.columns and rows_to_divide_new_df_indices:
                df.loc[rows_to_divide_new_df_indices, col] = (
                    pd.to_numeric(
                        df.loc[rows_to_divide_new_df_indices, col], errors="coerce"
                    )
                    / 1_000_000
                ).round(2)
            elif col not in df.columns:
                print(f"Warning: '{col}' column not found.")

        columns_to_format_indices_28_30 = (
            columns_to_divide_existing + ["Total"] + columns_to_divide_new
        )
        if 28 in df.index:
            for col in columns_to_format_indices_28_30:
                if col in df.columns:
                    val = pd.to_numeric(df.loc[28, col], errors="coerce")
                    df.loc[28, col] = round(val, 2) if pd.notna(val) else val
        if 29 in df.index:
            for col in columns_to_format_indices_28_30:
                if col in df.columns:
                    val = pd.to_numeric(df.loc[29, col], errors="coerce")
                    df.loc[29, col] = round(val, 1) if pd.notna(val) else val
        if 30 in df.index:
            for col in columns_to_format_indices_28_30:
                if col in df.columns:
                    val = pd.to_numeric(df.loc[30, col], errors="coerce")
                    df.loc[30, col] = round(val, 1) if pd.notna(val) else val

        if 31 in df.index:
            for col in columns_to_divide_index_31:
                if col in df.columns:
                    raw_val = pd.to_numeric(df.loc[31, col], errors="coerce")
                    if pd.notna(raw_val):
                        df.loc[31, col] = round(raw_val / 1_000_000, 2)

        # === Step 4.2: Add a blank column before "Đông Bắc Bộ" ===
        blank_col_inserted_name = "BLANK_COL_TEMP_NAME_FOR_LOC"
        if "Đông Bắc Bộ" in df.columns:
            blank_column_df_idx = df.columns.get_loc("Đông Bắc Bộ")
            df.insert(blank_column_df_idx, blank_col_inserted_name, "")
        else:
            print(
                "Warning: 'Đông Bắc Bộ' column not found. Skipping blank column insertion."
            )

        # === Step 5: Create Excel writer and write DataFrame ===
        writer = pd.ExcelWriter(output_file, engine="openpyxl")
        df.to_excel(writer, sheet_name="Report", index=False, startrow=1)

        workbook = writer.book
        worksheet = writer.sheets["Report"]

        # === Define cell styles ===
        header_font = Font(name="Calibri", size=12, bold=True, color="000000")
        cell_font = Font(name="Calibri", size=11, color="000000")
        header_fill = PatternFill(
            start_color="BDCFEF", end_color="BDCFEF", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="00FF00", end_color="00FF00", fill_type="solid"
        )
        gray_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )
        default_data_fill = PatternFill(
            start_color="FFFCC9", end_color="FFFCC9", fill_type="solid"
        )  # Updated default fill
        bold_first_col_font = Font(name="Calibri", size=11, bold=True, color="000000")
        highlight0_font = Font(name="Calibri", size=11, bold=True, color="000000")
        highlight0_fill = PatternFill(
            start_color="6495ED", end_color="6495ED", fill_type="solid"
        )
        fill_26 = PatternFill(
            start_color="E3B825", end_color="E3B825", fill_type="solid"
        )
        fill_27 = PatternFill(
            start_color="F07A17", end_color="F07A17", fill_type="solid"
        )
        bold_row_font = Font(name="Calibri", size=11, bold=True, color="000000")

        center_alignment = Alignment(horizontal="center", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        group_header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        group_header_fill = PatternFill(
            start_color="0000FF", end_color="0000FF", fill_type="solid"
        )
        special_rows_fill = PatternFill(
            start_color="A7FCF9", end_color="A7FCF9", fill_type="solid"
        )
        special_rows_number_format = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'

        # === Step 6: Format column headers (row 2) ===
        blank_col_excel_idx = -1
        if blank_col_inserted_name in df.columns:
            blank_col_excel_idx = df.columns.get_loc(blank_col_inserted_name) + 1

        for col_num, column_title_from_df in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=2, column=col_num)
            if column_title_from_df == "funding_name":
                cell.value = ""
            elif column_title_from_df == blank_col_inserted_name:
                cell.value = ""
            else:
                cell.value = column_title_from_df

            cell.font = header_font
            cell.alignment = center_alignment

            if column_title_from_df in ["TOTAL", "Month"]:
                cell.fill = gray_fill
            elif col_num == blank_col_excel_idx:
                cell.fill = yellow_fill
            else:
                cell.fill = header_fill

        # === Step 8: Format data cells (rows start at Excel row 3) ===
        special_indices_for_formatting_A = {1, 7, 12, 19, 20, 25}
        special_indices_for_formatting_B_N = {1, 7, 12, 19, 20, 25}

        for row_num_excel in range(3, len(df) + 3):
            df_index = row_num_excel - 3
            for col_num_excel in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num_excel, column=col_num_excel)
                actual_df_column_name = df.columns[col_num_excel - 1]

                # Apply default font and alignment
                cell.font = cell_font
                cell.alignment = left_alignment

                # Apply new default background fill (except first column)
                # This will be overridden by more specific fill rules below.
                if col_num_excel > 1:
                    cell.fill = default_data_fill

                # --- Special column styling for "TOTAL" and "Month" (Fill and Font) ---
                if actual_df_column_name == "TOTAL" or actual_df_column_name == "Month":
                    cell.fill = gray_fill  # Overrides default_data_fill
                    cell.font = bold_row_font
                    if actual_df_column_name == "TOTAL":
                        cell.alignment = right_alignment
                    # else "Month" keeps left_alignment (or could be center_alignment if preferred)

                # --- Rule 0: df_index == 0, Excel cols B-N (2-14) ---
                if df_index == 0 and 2 <= col_num_excel <= 14:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    cell.fill = PatternFill(
                        start_color="F68216", end_color="F68216", fill_type="solid"
                    )  # Overrides default/gray
                    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                    cell.alignment = right_alignment
                    if col_num_excel == blank_col_excel_idx:
                        cell.fill = yellow_fill  # Yellow override for blank
                    continue

                # --- Rule: Special indices (1,7,12,19,20,25), Excel cols B-N (2-14) ---
                elif (
                    df_index in special_indices_for_formatting_B_N
                    and 2 <= col_num_excel <= 14
                ):
                    cell.fill = special_rows_fill  # Overrides default/gray
                    cell.font = bold_row_font
                    cell.number_format = special_rows_number_format
                    cell.alignment = right_alignment
                    if col_num_excel == blank_col_excel_idx:
                        cell.fill = yellow_fill  # Yellow override for blank
                    continue

                # --- USER REQUESTED: df_index 4 OR (6-25), Excel cols H-N (8-14) ---
                elif (df_index == 4 or (6 <= df_index <= 25)) and (
                    8 <= col_num_excel <= 14
                ):
                    cell.number_format = special_rows_number_format
                    cell.alignment = right_alignment
                    # Font: If TOTAL/Month, it's already bold from gray_fill rule. Else, default cell_font.
                    if not (
                        actual_df_column_name == "TOTAL"
                        or actual_df_column_name == "Month"
                    ):
                        cell.font = cell_font
                    # Fill: default_data_fill or gray_fill is already set. Override if blank.
                    if col_num_excel == blank_col_excel_idx:
                        cell.fill = yellow_fill  # Overrides default/gray
                    continue

                # --- Rule: Col B (Excel col 2), if not caught by above specific B-N rules ---
                elif col_num_excel == 2:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    cell.alignment = right_alignment
                    if not (
                        actual_df_column_name == "TOTAL"
                        or actual_df_column_name == "Month"
                    ):  # Col B unlikely to be "Month"
                        cell.font = cell_font
                    if col_num_excel == blank_col_excel_idx:
                        cell.fill = yellow_fill
                    continue

                # --- Rule: Col A (Excel col 1), df_index == 0 ---
                elif (
                    df_index == 0 and col_num_excel == 1
                ):  # First column, default_data_fill was NOT applied
                    cell.font = highlight0_font
                    cell.fill = highlight0_fill  # Specific fill for this cell
                    cell.alignment = left_alignment
                    continue

                # --- Rule: Col A (Excel col 1), special_indices_for_formatting_A ---
                elif (
                    df_index in special_indices_for_formatting_A and col_num_excel == 1
                ):  # First column
                    cell.font = bold_first_col_font
                    cell.fill = green_fill  # Specific fill for these cells
                    cell.alignment = left_alignment
                    continue

                # --- General cell formatting if no 'continue' was hit ---
                # Font and alignment are set. Default fill might be set.
                # Specific fills for TOTAL/Month are set.

                # Ensure blank column is yellow if not handled by a more specific fill in a 'continue' block
                if col_num_excel == blank_col_excel_idx:
                    cell.fill = yellow_fill

                # --- Number formatting for remaining cells ---
                if actual_df_column_name == "Head":
                    if df_index in [28, 29, 30, 31]:
                        cell.number_format = numbers.FORMAT_GENERAL
                    else:
                        cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    cell.alignment = right_alignment
                elif actual_df_column_name == "Total" and df_index <= 25:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    cell.alignment = right_alignment
                elif actual_df_column_name in columns_to_divide_existing or (
                    actual_df_column_name == "Total" and df_index in [26, 27]
                ):
                    if actual_df_column_name == "Total":
                        if df_index == 26:
                            cell.number_format = (
                                '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                            )
                        elif df_index == 27:
                            cell.number_format = (
                                '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
                            )
                    elif actual_df_column_name in [
                        "Miền Bắc",
                        "Miền Nam",
                        "Miền Trung",
                    ]:
                        if df_index in [28, 29, 30, 31]:
                            cell.number_format = numbers.FORMAT_GENERAL
                        else:
                            cell.number_format = (
                                '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
                            )
                    cell.alignment = right_alignment
                elif actual_df_column_name == "TOTAL" and df_index in [28, 29, 30]:
                    cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    # cell.alignment = right_alignment # Already set by TOTAL/Month rule
                elif (
                    actual_df_column_name in columns_to_divide_new
                    and df_index not in [26, 28, 29, 30, 31]
                ):
                    if actual_df_column_name == "TOTAL":
                        cell.number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
                    else:
                        cell.number_format = (
                            '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
                        )
                    cell.alignment = right_alignment
                elif (
                    df_index == 28
                    and actual_df_column_name in columns_to_format_indices_28_30
                ):
                    cell.number_format = (
                        '_(* #,##0.00_);_(* -#,##0.00_);_(* "-"_);_(@_)'
                    )
                    cell.alignment = right_alignment
                elif (
                    df_index == 29
                    and actual_df_column_name in columns_to_format_indices_28_30
                ):
                    cell.number_format = '_(* #,##0.0_);_(* -#,##0.0_);_(* "-"_);_(@_)'
                    cell.alignment = right_alignment
                elif (
                    df_index == 30
                    and actual_df_column_name in columns_to_format_indices_28_30
                ):
                    cell.number_format = '_(* #,##0.0_);_(* -#,##0.0_);_(* "-"_);_(@_)'
                    cell.alignment = right_alignment
                elif (
                    df_index == 31
                    and actual_df_column_name in columns_to_divide_index_31
                ):
                    cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"_);_(@_)'
                    cell.alignment = right_alignment

                # --- Row-specific fill/font overrides (apply last for these rows if conditions met) ---
                # These can override any previous fill (default_data_fill, gray_fill, yellow_fill etc.)
                if df_index == 26:
                    if not (
                        2 <= col_num_excel <= 5 or col_num_excel == blank_col_excel_idx
                    ):
                        cell.fill = fill_26
                        cell.font = bold_row_font
                elif df_index == 27:
                    if not (
                        2 <= col_num_excel <= 5 or col_num_excel == blank_col_excel_idx
                    ):
                        cell.fill = fill_27
                        cell.font = bold_row_font

        # === Step 9: Adjust column widths ===
        for col_idx_for_width, df_col_name_for_width in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx_for_width)
            if col_idx_for_width == blank_col_excel_idx:
                worksheet.column_dimensions[column_letter].width = 3
            else:
                max_len = 0
                header_cell_val = worksheet.cell(row=2, column=col_idx_for_width).value
                if header_cell_val:
                    max_len = len(str(header_cell_val))

                for data_val in df[df_col_name_for_width]:
                    if pd.notnull(data_val):
                        max_len = max(max_len, len(str(data_val)))

                adj_width = max_len + 2 if max_len > 0 else 10
                worksheet.column_dimensions[column_letter].width = min(adj_width, 50)

        # === Step 10: Set row heights ===
        worksheet.row_dimensions[1].height = 20
        worksheet.row_dimensions[2].height = 30
        for i in range(3, len(df) + 3):
            worksheet.row_dimensions[i].height = 20

        # === Step 11: Add merged headers ===
        col_head_idx = df.columns.get_loc("Head") + 1 if "Head" in df.columns else -1
        col_total_idx = df.columns.get_loc("Total") + 1 if "Total" in df.columns else -1
        if col_head_idx != -1 and col_total_idx != -1 and col_head_idx <= col_total_idx:
            cell_mg1 = worksheet.cell(row=1, column=col_head_idx)
            cell_mg1.value = "Tổng cần phân bổ xuống cho ĐVML"
            cell_mg1.font = group_header_font
            cell_mg1.fill = group_header_fill
            cell_mg1.alignment = center_alignment
            worksheet.merge_cells(
                start_row=1,
                start_column=col_head_idx,
                end_row=1,
                end_column=col_total_idx,
            )

        col_dbb_idx = (
            df.columns.get_loc("Đông Bắc Bộ") + 1 if "Đông Bắc Bộ" in df.columns else -1
        )
        col_kvml_total_idx = (
            df.columns.get_loc("TOTAL") + 1 if "TOTAL" in df.columns else -1
        )
        if (
            col_dbb_idx != -1
            and col_kvml_total_idx != -1
            and col_dbb_idx <= col_kvml_total_idx
        ):
            cell_mg2 = worksheet.cell(row=1, column=col_dbb_idx)
            cell_mg2.value = "KHU VỰC MẠNG LƯỚI"
            cell_mg2.font = group_header_font
            cell_mg2.fill = group_header_fill
            cell_mg2.alignment = center_alignment
            worksheet.merge_cells(
                start_row=1,
                start_column=col_dbb_idx,
                end_row=1,
                end_column=col_kvml_total_idx,
            )

        # === Step 12: Save the file locally ===
        writer.close()
        writer = None
        print(f"Excel file created successfully: {output_file}")

        # === Step 13: Upload to Google Drive ===
        print("\n--- INITIATING GOOGLE DRIVE UPLOAD ---")
        print("Authenticating...")
        drive_service = authenticate_google_drive()
        if drive_service:
            print(f"Uploading file '{output_file}'...")
            upload_to_drive(drive_service, output_file)

    except Error as db_error:
        print(f"Database Error: {db_error}")
    except Exception as error:
        print(f"General Error: {error}")
        import traceback

        traceback.print_exc()
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()
            print("PostgreSQL connection closed.")
        if writer is not None:
            writer.close()


if __name__ == "__main__":
    db_params = {
        "host": "localhost",
        "port": "5432",
        "dbname": "final_project",
        "user": "postgres",
        "password": "1234",
    }
    query = """
    SELECT 
        d.funding_name, 
        f.tpb_head AS "Head", f.tpb_mienbac AS "Miền Bắc", f.tpb_miennam AS "Miền Nam",
        f.tpb_mientrung AS "Miền Trung", f.tpv_total AS "Total", 
        f.kvml_dbb AS "Đông Bắc Bộ", f.kvml_tbb AS "Tây Bắc Bộ", f.kvml_dbsh AS "ĐB Sông Hồng",
        f.kvml_btb AS "Bắc Trung Bộ", f.kvml_ntb AS "Nam Trung Bộ", f.kvml_tnb AS "Tây Nam Bộ",
        f.kvml_dnb AS "Đông Nam Bộ", f.kvml_total AS "TOTAL", f.month_key AS "Month"
    FROM dim_funding_structure d 
    JOIN fact_backdate_funding_monthly f ON d.funding_id = f.funding_id 
    WHERE f.month_key = 202302
    ORDER BY d.sortorder;
    """
    output_file = "BaocaoTonghop_T2_formatted_py.xlsx"  # Changed output filename
    export_postgres_to_excel(db_params, query, output_file)
