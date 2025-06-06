import psycopg2
import pandas as pd
from psycopg2 import Error
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload

# --- Google Drive Configuration ---
# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/drive.file"]


def authenticate_google_drive():
    """Authenticates with the Google Drive API and returns the service object."""
    creds = None

    # Get the absolute path of the directory containing the running script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    token_path = os.path.join(script_dir, "token.json")
    credential_path = os.path.join(script_dir, "credentials.json")

    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first time.
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)

    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # --- PATH FIX: Use absolute path to credentials.json ---
            print(f"Searching for credentials file at: {credential_path}")
            if not os.path.exists(credential_path):
                print("\nERROR: Could not find credentials.json at the specified path.")
                print(">>> Please CHECK AGAIN and ensure that:")
                print(
                    "1. The file is named exactly 'credentials.json' (not 'credentials.json.txt')."
                )
                print(
                    "2. The 'credentials.json' file is placed in the SAME FOLDER as this Python script.\n"
                )
                return None

            flow = InstalledAppFlow.from_client_secrets_file(credential_path, SCOPES)
            creds = flow.run_local_server(port=0)

        # Save the credentials for the next run
        with open(token_path, "w") as token:
            token.write(creds.to_json())

    try:
        service = build("drive", "v3", credentials=creds)
        print("Google Drive authentication successful.")
        return service
    except HttpError as error:
        print(f"An error occurred while building the Drive service: {error}")
        return None


def upload_to_drive(service, file_path, folder_id=None):
    """Uploads a file to Google Drive, optionally to a specific folder."""
    if not service:
        print("Upload failed: Google Drive service is not authenticated.")
        return

    file_name = os.path.basename(file_path)
    # Define file metadata, including the parent folder if an ID is provided
    file_metadata = {"name": file_name}
    if folder_id:
        file_metadata["parents"] = [folder_id]

    media = MediaFileUpload(
        file_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    try:
        file = (
            service.files()
            .create(body=file_metadata, media_body=media, fields="id, webViewLink")
            .execute()
        )
        print(f"File upload successful! File ID: {file.get('id')}")
        print(f"You can view the file at: {file.get('webViewLink')}")
    except HttpError as error:
        print(f"An error occurred during file upload: {error}")


def export_postgres_to_excel(db_params, query, output_file):
    """
    Export data from PostgreSQL to an Excel file with custom formatting.

    Parameters:
    - db_params (dict): Database connection parameters (host, port, dbname, user, password).
    - query (str): SQL query to fetch data.
    - output_file (str): Path to the output Excel file.
    """
    connection = None
    try:
        # Step 1: Connect to PostgreSQL
        print("Connecting to PostgreSQL...")
        connection = psycopg2.connect(
            host=db_params["host"],
            port=db_params["port"],
            database=db_params["dbname"],
            user=db_params["user"],
            password=db_params["password"],
        )
        cursor = connection.cursor()

        # Step 2: Execute the query
        print("Executing query...")
        cursor.execute(query)

        # Step 3: Fetch column names and data
        columns = [desc[0] for desc in cursor.description]
        data = cursor.fetchall()

        # Step 4: Create a pandas DataFrame
        print("Loading data into DataFrame...")
        df = pd.DataFrame(data, columns=columns)

        # Convert tongdiem to integer type
        if "Tổng điểm" in df.columns:
            df["Tổng điểm"] = df["Tổng điểm"].astype(int)

        if "psdn_avg" in df.columns:
            df["psdn_avg"] = (
                df["psdn_avg"]
                .astype(float)
                .apply(
                    lambda x: int(round(x, 1))
                    if round(x, 1).is_integer()
                    else round(x, 1)
                )
            )

        for col in ["ltn_avg", "hsbq_nhan_su"]:
            if col in df.columns:
                df[col] = (
                    df[col]
                    .astype(float)
                    .apply(
                        lambda x: int(round(x, 2))
                        if round(x, 2).is_integer()
                        else round(x, 2)
                    )
                )

        for col in [
            "cir",
            "margin",
            "hs_von",
            "approval_rate_avg",
            "npl_truoc_wo_luy_ke",
        ]:
            if col in df.columns:
                df[col] = df[col].astype(float)

        # Step 5: Create Excel writer
        writer = pd.ExcelWriter(output_file, engine="openpyxl")
        df.to_excel(
            writer, sheet_name="Report", index=False, startrow=1
        )  # Start from row 2 for header

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets["Report"]

        # Define styles
        header_font = Font(name="Calibri", size=12, bold=True)
        cell_font = Font(name="Calibri", size=11)
        cell_bold_font = Font(
            name="Calibri", size=11, bold=True
        )  # Bold font for data cells
        header_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
        )  # Light blue for column headers
        alternate_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        )  # Light gray
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        left_alignment = Alignment(horizontal="left", vertical="center")
        # New fills for Quy Mô and Tài Chính
        quy_mo_fill = PatternFill(
            start_color="0000FF", end_color="0000FF", fill_type="solid"
        )  # Blue
        tai_chinh_fill = PatternFill(
            start_color="FFA500", end_color="FFA500", fill_type="solid"
        )  # Orange
        white_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

        # New fill for special headers (Green)
        green_fill = PatternFill(
            start_color="00AA00", end_color="00AA00", fill_type="solid"
        )

        # New fills for additional custom formatting
        light_green_fill = PatternFill(
            start_color="90EE90", end_color="90EE90", fill_type="solid"
        )  # Light green for "Tổng điểm"
        blue_fill = PatternFill(
            start_color="0070C0", end_color="0070C0", fill_type="solid"
        )  # Blue for "rank_final"
        light_blue_fill = PatternFill(
            start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"
        )  # Light blue for "Điểm Quy Mô" and "Điểm FIN"

        # Step 6: Format the column headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=2, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.border = border
            # Apply right alignment for specific columns
            if column_title in ["Tổng điểm", "Điểm Quy Mô", "Điểm FIN"]:
                cell.alignment = right_alignment
            else:
                cell.alignment = center_alignment

            # Apply special formatting based on column title
            if column_title in ["rank_ptkd", "rank_fin"]:
                cell.fill = green_fill
                cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            elif column_title == "Tổng điểm":
                cell.fill = light_green_fill
                cell.font = Font(name="Calibri", size=12, bold=True)  # Ensure bold
            elif column_title == "rank_final":
                cell.fill = blue_fill
                cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            elif column_title in ["Điểm Quy Mô", "Điểm FIN"]:
                cell.fill = light_blue_fill
            else:
                cell.fill = header_fill

        # Step 7: Format data rows
        for row_num in range(3, len(df) + 3):  # Data starts from row 3
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                # Apply bold font for specific columns
                if df.columns[col_num - 1] in [
                    "Tổng điểm",
                    "rank_final",
                    "rank_ptkd",
                    "rank_fin",
                ]:
                    cell.font = cell_bold_font
                else:
                    cell.font = cell_font
                cell.border = border
                # Apply alternate row colors
                if row_num % 2 == 0:
                    cell.fill = alternate_fill
                # Align numeric columns to right, others to left
                if df.columns[col_num - 1] in [
                    "Tổng điểm",
                    "rank_final",
                    "ltn_avg",
                    "rank_ltn_avg",
                    "psdn_avg",
                    "rank_psdn_avg",
                    "approval_rate_avg",
                    "rank_approval_rate_avg",
                    "npl_truoc_wo_luy_ke",
                    "rank_npl_truoc_wo_luy_ke",
                    "Điểm Quy Mô",
                    "rank_ptkd",
                    "cir",
                    "rank_cir",
                    "margin",
                    "rank_margin",
                    "hs_von",
                    "rank_hs_von",
                    "hsbq_nhan_su",
                    "rank_hsbq_nhan_su",
                    "Điểm FIN",
                    "rank_fin",
                ]:
                    cell.alignment = right_alignment
                else:
                    cell.alignment = left_alignment
                # Format numbers
                if df.columns[col_num - 1] in ["ltn_avg", "hsbq_nhan_su"]:
                    cell.number_format = (
                        "#,##0" if isinstance(cell.value, int) else "#,##0.00"
                    )
                elif df.columns[col_num - 1] == "psdn_avg":
                    cell.number_format = (
                        "#,##0" if isinstance(cell.value, int) else "#,##0.0"
                    )

        # Step 8: Adjust column widths
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                max((len(str(val)) for val in df[column]), default=10),
                len(column),
            )
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Step 9: Set row heights
        worksheet.row_dimensions[1].height = 20  # Group header row
        worksheet.row_dimensions[2].height = 30  # Column header row
        for row_num in range(3, len(df) + 3):
            worksheet.row_dimensions[row_num].height = 20  # Data rows

        # Step 10: Add merged column headers
        # Merged column header for "Quy Mô"
        quy_mo_cell = worksheet.cell(row=1, column=6)
        quy_mo_cell.value = "QUY MÔ"
        quy_mo_cell.font = white_font
        quy_mo_cell.alignment = center_alignment
        quy_mo_cell.fill = quy_mo_fill
        quy_mo_cell.border = border
        worksheet.merge_cells(start_row=1, start_column=6, end_row=1, end_column=15)

        # Merged column header for "Tài Chính"
        tai_chinh_cell = worksheet.cell(row=1, column=16)
        tai_chinh_cell.value = "TÀI CHÍNH"
        tai_chinh_cell.font = white_font
        tai_chinh_cell.alignment = center_alignment
        tai_chinh_cell.fill = tai_chinh_fill
        tai_chinh_cell.border = border
        worksheet.merge_cells(start_row=1, start_column=16, end_row=1, end_column=25)

        # Step 11: Save the file
        writer.close()
        writer = None
        print(f"Excel file created successfully: {output_file}")

        # === Step 12: Upload to Google Drive ===
        print("\n--- INITIATING GOOGLE DRIVE UPLOAD ---")
        print("Authenticating...")
        drive_service = authenticate_google_drive()
        if drive_service:
            print(f"Uploading file '{output_file}'...")
            upload_to_drive(drive_service, output_file)

    except (Exception, Error) as error:
        print(f"Error: {error}")

    finally:
        # Step 12: Close database connection
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection closed.")


# Example usage
if __name__ == "__main__":
    # Database connection parameters
    db_params = {
        "host": "localhost",
        "port": "5432",
        "dbname": "final_project",
        "user": "postgres",
        "password": "1234",
    }

    # SQL query
    query = """
    SELECT 
        f.month_key 
        , f.area_cde 
        , f.email 
        , f.tongdiem AS "Tổng điểm"
        , f.rank_final 
        , f.ltn_avg 
        , f.rank_ltn_avg 
        , f.psdn_avg 
        , f.rank_psdn_avg 
        , f.approval_rate_avg 
        , f.rank_approval_rate_avg 
        , f.npl_truoc_wo_luy_ke 
        , f.rank_npl_truoc_wo_luy_ke 
        , f.diem_quy_mo AS "Điểm Quy Mô"
        , f.rank_ptkd 
        , f.cir 
        , f.rank_cir 
        , f.margin 
        , f.rank_margin 
        , f.hs_von 
        , f.rank_hs_von 
        , f.hsbq_nhan_su 
        , f.rank_hsbq_nhan_su 
        , f.diem_fin AS "Điểm FIN"
        , f.rank_fin
    FROM fact_backdate_asm_monthly f
    WHERE month_key = 202302
    ORDER BY f.rank_final ;
    """

    # Output Excel file path
    output_file = "BaocaoXepHangASM_T2_formatted_py.xlsx"

    # Call the function
    export_postgres_to_excel(db_params, query, output_file)
